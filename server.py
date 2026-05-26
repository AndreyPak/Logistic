#!/usr/bin/env python3
"""
Локальный сервер для Olive Logistics.
Запуск: python3 server.py
Открыть: http://localhost:8765
"""
from http.server import HTTPServer, BaseHTTPRequestHandler
from datetime import datetime
import os, re, io

# ─── Путь к папке Output ────────────────────────────────────────────────────
BASE = os.path.dirname(os.path.abspath(__file__))

# ─── Весь код генерации Excel ────────────────────────────────────────────────
def build_excel():
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Border, Side

    PRODS = [{"type": "category", "name": "Завтрак 250-350"}, {"type": "product", "artikul": 25326, "name": "ПП* Упак Банановый кекс (1шт)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 29718, "name": "ПП* Упак Домашний творог с лесными ягодами (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 19176, "name": "ПП* Упак Каша из полбы с клюквой (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 376, "name": "ПП* Упак Каша овсяная с фруктами (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 19180, "name": "ПП* Упак Нежная кукурузная каша с джемом из кураги и семенами чиа (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 98693, "name": "ПП* Упак Омлет с грибами (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 19291, "name": "ПП* Упак Сырники с чиа (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 59883, "name": "ПП* Упак Чиа пуддинг манго (1порц)", "kratnost": "1 порц", "itogo": 1}, {"type": "product", "artikul": 695, "name": "ПП* Упак Сэндвич с индейкой (0,5 шт)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 697, "name": "ПП* Упак Сэндвич с лососем (0,5 шт)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 693, "name": "ПП* Упак Сэндвич с ростбифом (0,5 шт)", "kratnost": "1 порц", "itogo": 0}, {"type": "category", "name": "Завтрак 400-500"}, {"type": "product", "artikul": 681, "name": "ПП* Упак Блины с ветчиной и сыром (3шт)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 98, "name": "ПП* Упак Живая гречка с семенами ,гранолой и чиа (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 91716, "name": "ПП* Упак Завтрак с глазуньей и креветками (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 88807, "name": "ПП* Упак Завтрак с лососем и глазуньей (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 88711, "name": "ПП* Упак Йогурт с гранолой и курагой (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 88701, "name": "ПП* Упак Йогурт с гранолой и малиной (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 405, "name": "ПП* Упак Кесадилья с яичным паштетом (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 19285, "name": "ПП* Упак Омлет с ветчиной (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 19284, "name": "ПП* Упак Омлет соленый лосось (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 387, "name": "ПП* Упак Ролл-салат по азиатски (1шт)", "kratnost": "1 порц", "itogo": 1}, {"type": "product", "artikul": 29270, "name": "ПП* Упак Ролл-салат с курицей (1шт)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 88708, "name": "ПП* Упак Скрембл с индейкой (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 90127, "name": "ПП* Упак Сэндвич с лососем (1шт)", "kratnost": "1 порц", "itogo": 2}, {"type": "product", "artikul": 411, "name": "ПП* Упак Шпинатные блинчики с курицей (2шт)", "kratnost": "1 порц", "itogo": 0}, {"type": "category", "name": "Горячее 400-500"}, {"type": "product", "artikul": 352, "name": "ПП* Упак Белая рыба со шпинатом и пюре с цветной капусты (1порц)", "kratnost": "1 порц", "itogo": 4}, {"type": "product", "artikul": 381, "name": "ПП* Упак Бигус (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 88848, "name": "ПП* Упак Индейка сувид с перловкой и соусом зеленым велюте (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 369, "name": "ПП* Упак Куриная грудинка с запеченным картофелем (1порц)", "kratnost": "1 порц", "itogo": 5}, {"type": "product", "artikul": 480, "name": "ПП* Упак Куриная грудка с нутом (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 55979, "name": "ПП* Упак Куриная грудка с орзо (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 378, "name": "ПП* Упак Куриное бедро с рагу из овощей (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 88825, "name": "ПП* Упак Куриное филе с киноа и соусом грин (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 88881, "name": "ПП* Упак Куриное филе с рисом и зеленым соусом велюте (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 88878, "name": "ПП* Упак Куриный Кебаб с овощами гриль (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 19226, "name": "ПП* Упак Курица с овощами на гриле (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 19229, "name": "ПП* Упак Лапша соба с говядиной (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 19230, "name": "ПП* Упак Лапша соба с курицей (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 366, "name": "ПП* Упак Люля куриный с овощами (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 57926, "name": "ПП* Упак Мясо по-казахски с картофелем и семолиной (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 91735, "name": "ПП* Упак Паста цельнозерновая с курицей сувид (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 34380, "name": "ПП* Упак Поке с креветками (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 37378, "name": "ПП* Упак Поке с курицей (1шт)", "kratnost": "1 порц", "itogo": 1}, {"type": "product", "artikul": 94379, "name": "ПП* Упак Поке с лососем (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 88876, "name": "ПП* Упак Семга с бурым рисом (1порц)", "kratnost": "1 порц", "itogo": 1}, {"type": "product", "artikul": 88852, "name": "ПП* Упак Сочные митболы с картофельным пюре и овощами (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 39219, "name": "ПП* Упак Танкацу с брокколи и картофельным пюре (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 61720, "name": "ПП* Упак Томленный гуляш с польбой и овощами (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 88817, "name": "ПП* Упак Тучикены с соусом арабиата и гречкой 1 пор", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 31095, "name": "ПП* Упак Фит панкейки из кабачков с соусом дзадзыки (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "category", "name": "Горячее 500-600"}, {"type": "product", "artikul": 348, "name": "ПП* Упак Белая рыба со шпинатом и картофельным пюре (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 88875, "name": "ПП* Упак Бифштекс с яйцом и картофелем (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 88832, "name": "ПП* Упак Индейка сувид с гречкой и грибным соусом (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 19219, "name": "ПП* Упак Котлета рыбная с рисом и овощами (1порц)", "kratnost": "1 порц", "itogo": 1}, {"type": "product", "artikul": 478, "name": "ПП* Упак Куриная грудка с чечевицей (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 66984, "name": "ПП* Упак Паста болоньезе (1порц)", "kratnost": "1 порц", "itogo": 1}, {"type": "product", "artikul": 66983, "name": "ПП* Упак Паста с курицей и грибами (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "category", "name": "Суп 200"}, {"type": "product", "artikul": 88860, "name": "ПП* Упак Крем-суп из запеченной тыквы с чечевицей (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 44928, "name": "ПП* Упак Суп борщ с говядиной (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 19249, "name": "ПП* Упак Суп тыквенный (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 18250, "name": "ПП* Упак Суп щи с кислой капустой (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 88868, "name": "ПП* Упак Тайский суп (1порц)", "kratnost": "1 порц", "itogo": 1}, {"type": "product", "artikul": 19253, "name": "ПП* Упак Харчо с насыщенным пряным бульоном (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "category", "name": "Салат 150-250"}, {"type": "product", "artikul": 37967, "name": "ПП* Упак Грин детокс салат с брокколи и семенами тыквы (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 484, "name": "ПП* Упак Грин микс с огурцом и горошком (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 52968, "name": "ПП* Упак Зеленый салат с брокколи и цитрусами (1порц)", "kratnost": "1 порц", "itogo": 6}, {"type": "product", "artikul": 52969, "name": "ПП* Упак Зеленый салат с оливками (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 78940, "name": "ПП* Упак Классический греческий салат с фетой и оливковым маслом (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 52970, "name": "ПП* Упак Миск зеленых овощей с авокадо и томатами черри (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 57930, "name": "ПП* Упак Салат Витаминка (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 57371, "name": "ПП* Упак Салат с индейкой (1порц)", "kratnost": "1 порц", "itogo": 1}, {"type": "product", "artikul": 488, "name": "ПП* Упак Шпинатный салат с черри и маслинами (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "category", "name": "Салат 250-350"}, {"type": "product", "artikul": 100, "name": "ПП* Упак Салат Детокс (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 85219, "name": "ПП* Упак Салат итальянский с соусом песто (1порц)", "kratnost": "1 порц", "itogo": 1}, {"type": "product", "artikul": 85221, "name": "ПП* Упак Салат с лососем (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 64386, "name": "ПП* Упак Салат с тыквой и киноа (1порц)", "kratnost": "1 порц", "itogo": 1}, {"type": "product", "artikul": 85220, "name": "ПП* Упак Салат Цезарь с курицей (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "category", "name": "Выпечка/Десерт 100-250"}, {"type": "product", "artikul": 17857, "name": "Овсяные хлебцы", "kratnost": "1 порц", "itogo": 0}, {"type": "category", "name": "Выпечка/Десерт 300-350"}, {"type": "category", "name": "Смузи 100-150"}, {"type": "product", "artikul": 47537, "name": "ПП* Упак Смузи морковь,апельсин,тыква (1шт)", "kratnost": "1 порц", "itogo": 0}, {"type": "category", "name": "Сендвичи 300-350"}, {"type": "category", "name": "Блюда на замену"}, {"type": "product", "artikul": 405, "name": "ПП* Упак Кесадилья с яичным паштетом (1порц)", "kratnost": "1 порц", "itogo": 3}, {"type": "product", "artikul": 19281, "name": "ПП* Упак Омлет запеченная курица (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 130, "name": "ПП* Упак Сэндвич с ростбифом (1шт)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 414, "name": "ПП* Упак Говядина с тушеными овощами и булгуром (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 19224, "name": "ПП* Упак Жаренный лосось с азиатским рисом (1порц)", "kratnost": "1 порц", "itogo": 3}, {"type": "product", "artikul": 88825, "name": "ПП* Упак Куриное филе с киноа и соусом грин (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 88837, "name": "ПП* Упак Пенне из 5 злаков с куриным филе под соусом зеленый велюте (1порц)", "kratnost": "1 порц", "itogo": 1}, {"type": "product", "artikul": 91716, "name": "ПП* Упак Завтрак с глазуньей и креветками (1порц)", "kratnost": "1 порц", "itogo": 2}, {"type": "product", "artikul": 19284, "name": "ПП* Упак Омлет соленый лосось (1порц)", "kratnost": "1 порц", "itogo": 1}, {"type": "product", "artikul": 88885, "name": "ПП* Упак Нежная семга с овощами гриль (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 46945, "name": "ПП* Упак Паста с креветками (1порц)", "kratnost": "1 порц", "itogo": 1}, {"type": "product", "artikul": 361, "name": "ПП* Упак Ростбиф с гречкой и чими чури (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 358, "name": "ПП* Упак Ростбиф с грибами (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "product", "artikul": 517, "name": "ПП* Упак Спагетти с морепродуктами в томатном соусе (1порц)", "kratnost": "1 порц", "itogo": 0}, {"type": "category", "name": "Материалы"}, {"type": "product", "artikul": 27602, "name": "У* Термопакет серый 28*28*20", "kratnost": "1 шт", "itogo": 6}, {"type": "product", "artikul": 27605, "name": "У* Термопакет оранжевый 30*30*20", "kratnost": "1 шт", "itogo": 2}, {"type": "product", "artikul": 27603, "name": "У* Термопакет зеленый 28*28*20", "kratnost": "1 шт", "itogo": 1}, {"type": "product", "artikul": 27604, "name": "У* Термопакет красный 30*30*20", "kratnost": "1 шт", "itogo": 0}, {"type": "product", "artikul": 99727, "name": "У* Набор одноразовый (вилка,ложка,салфетка)", "kratnost": "1 шт", "itogo": 9}, {"type": "product", "artikul": 31880, "name": "С* Вода Тассай без газа 777 мл", "kratnost": "1 шт", "itogo": 9}, {"type": "product", "artikul": "", "name": "Хладогент", "kratnost": "1 шт", "itogo": 0}, {"type": "category", "name": "Полуфабрикаты для дома"}]

    LABELS = [{"id": "23883", "name": "Олеся", "phone4": "9458", "address": "Жансугурова, 176а/2, кв. -, домофон: -", "ration": "1200 ккал", "menu": [{"type": "Завтрак", "dish": "Домашний творог с лесными ягодами", "bold": False}, {"type": "Обед", "dish": "Грин детокс салат с брокколи и семенами тыквы", "bold": False}, {"type": "Обед", "dish": "Спагетти с морепродуктами в томатном соусе", "bold": True}, {"type": "Полдник", "dish": "Овсяные хлебцы", "bold": False}, {"type": "Ужин", "dish": "Бигус", "bold": False}]}, {"id": "23856", "name": "Айгерим", "phone4": "4911", "address": "Аль-Фараби, 41/3, кв. 124", "ration": "1800 ккал", "menu": [{"type": "Завтрак", "dish": "Омлет запеченная курица", "bold": True}, {"type": "Завтрак", "dish": "Салат с лососем", "bold": False}, {"type": "Обед", "dish": "Лапша соба с курицей", "bold": False}, {"type": "Обед", "dish": "Крем-суп из запеченной тыквы с чечевицей", "bold": False}, {"type": "Ужин", "dish": "Паста цельнозерновая с курицей сувид", "bold": False}]}, {"id": "23369", "name": "Андрей", "phone4": "0705", "address": "KOKTEM 3, 24, кв. 140, домофон: 140", "ration": "1800 ккал", "menu": []}, {"id": "23567", "name": "Алина Олив", "phone4": "8682", "address": "Кокшокы, 6/1, кв. 1, домофон: 1", "ration": "1500 ккал", "menu": [{"type": "Завтрак", "dish": "Скрембл с индейкой", "bold": False}, {"type": "Обед", "dish": "Салат детокс", "bold": False}, {"type": "Обед", "dish": "Бифштекс с яйцом и картофелем", "bold": False}, {"type": "Ужин", "dish": "Куриная грудка с запечённым картофелем", "bold": False}]}, {"id": "23569", "name": "Данияр Olive", "phone4": "9892", "address": "райымбека, 210/3 блок 11, кв. 260, домофон: 260", "ration": "1200 ккал", "menu": [{"type": "Завтрак", "dish": "Домашний творог с лесными ягодами", "bold": False}, {"type": "Обед", "dish": "Грин детокс салат с брокколи и семенами тыквы", "bold": False}, {"type": "Обед", "dish": "Индейка сувид с перловкой и соусом зеленым велюте", "bold": False}, {"type": "Полдник", "dish": "Овсяные хлебцы", "bold": False}, {"type": "Ужин", "dish": "Бигус", "bold": False}]}, {"id": "23615", "name": "Алексей", "phone4": "2329", "address": "СарыАрка, 1/4, 9, кв. 43, домофон: 43", "ration": "1200 ккал", "menu": [{"type": "Завтрак", "dish": "Кесадилья с яичным паштетом", "bold": True}, {"type": "Обед", "dish": "Грин детокс салат с брокколи и семенами тыквы", "bold": False}, {"type": "Обед", "dish": "Куриное филе с киноа и соусом грин", "bold": True}, {"type": "Полдник", "dish": "Овсяные хлебцы", "bold": False}, {"type": "Ужин", "dish": "Бигус", "bold": False}]}, {"id": "23683", "name": "Рустам Олив", "phone4": "9947", "address": "Навои, 37, кв. 332, домофон: 332", "ration": "1200 ккал", "menu": [{"type": "Завтрак", "dish": "Омлет запеченная курица", "bold": True}, {"type": "Обед", "dish": "Грин детокс салат с брокколи и семенами тыквы", "bold": False}, {"type": "Обед", "dish": "Индейка сувид с перловкой и соусом зеленым велюте", "bold": False}, {"type": "Полдник", "dish": "Овсяные хлебцы", "bold": False}, {"type": "Ужин", "dish": "Бигус", "bold": False}]}, {"id": "23739", "name": "Перизат", "phone4": "1444", "address": "Назарбаева 28а, 19, кв. 32, домофон: 32", "ration": "1200 ккал", "menu": [{"type": "Завтрак", "dish": "Завтрак с глазуньей и креветками", "bold": True}, {"type": "Обед", "dish": "Грин детокс салат с брокколи и семенами тыквы", "bold": False}, {"type": "Обед", "dish": "Сэндвич мини с ростбифом", "bold": True}, {"type": "Полдник", "dish": "Овсяные хлебцы", "bold": False}, {"type": "Ужин", "dish": "Бигус", "bold": False}]}, {"id": "23813", "name": "Алтынай", "phone4": "0714", "address": "Аль-Фараби, 34", "ration": "1200 ккал", "menu": [{"type": "Завтрак", "dish": "Домашний творог с лесными ягодами", "bold": False}, {"type": "Обед", "dish": "Грин детокс салат с брокколи и семенами тыквы", "bold": False}, {"type": "Обед", "dish": "Индейка сувид с перловкой и соусом зеленым велюте", "bold": False}, {"type": "Полдник", "dish": "Овсяные хлебцы", "bold": False}, {"type": "Ужин", "dish": "Бигус", "bold": False}]}]

    thin = Side(style='thin', color='000000')
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_hdr = PatternFill('solid', fgColor='D9D9D9')
    fill_sep = PatternFill('solid', fgColor='A6A6A6')
    fill_cat = PatternFill('solid', fgColor='FFFF00')
    fill_mat = PatternFill('solid', fgColor='8EA9DB')
    font_bold = Font(bold=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Прайс Сегодня'

    ws['A1'] = 1;          ws['B1'] = 'Заявочный лист "Фудзавод"'
    ws['A2'] = 'Информация заказчика'
    ws['A3'] = 'Дата';     ws['B3'] = datetime.now().strftime('%d.%m.%Y')
    ws['A4'] = 'Компания'; ws['B4'] = 'O-Live'
    ws['A5'] = 'Эл.почта'
    ws['A6'] = 'Конт.тел'
    ws['A7'] = 'Заполнил (И.Ф.)'

    for col, val in enumerate(['Артикул', 'Наименование продукта', 'Кратность заказа', 'ИТОГО'], 1):
        c = ws.cell(row=8, column=col, value=val)
        c.fill = fill_hdr; c.font = font_bold; c.border = border_all

    for col in range(1, 5):
        c = ws.cell(row=9, column=col)
        c.fill = fill_sep; c.border = border_all

    r = 10
    for p in PRODS:
        if p['type'] == 'category':
            fill = fill_mat if p['name'] in ('Материалы', 'Полуфабрикаты для дома') else fill_cat
            for col in range(1, 5):
                c = ws.cell(row=r, column=col); c.fill = fill; c.border = border_all
            ws.cell(row=r, column=1).value = p['name']
            ws.cell(row=r, column=1).font = font_bold
        else:
            ws.cell(row=r, column=1, value=p.get('artikul', '') or '').border = border_all
            ws.cell(row=r, column=2, value=p['name']).border = border_all
            ws.cell(row=r, column=3, value=p.get('kratnost', '')).border = border_all
            itogo = p.get('itogo', 0)
            ws.cell(row=r, column=4, value=itogo if itogo else '').border = border_all
        r += 1

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 10

    ws2 = wb.create_sheet('Этикетки')
    er = 1
    for lb in LABELS:
        ws2.cell(row=er, column=1, value='ФИО:').font = font_bold
        ws2.cell(row=er, column=2, value=lb['name'] + ' ' + lb['phone4']); er += 1
        ws2.cell(row=er, column=1, value='Адрес:').font = font_bold
        ws2.cell(row=er, column=2, value=lb['address']); er += 1
        er += 1
        ws2.cell(row=er, column=1, value='Рацион:').font = font_bold
        ws2.cell(row=er, column=2, value=lb['ration']).font = font_bold; er += 1
        er += 1
        ws2.cell(row=er, column=1, value='Меню').font = font_bold; er += 1
        for item in lb['menu']:
            c = ws2.cell(row=er, column=1, value=item['type'] + ': ' + item['dish'])
            if item['bold']: c.font = font_bold
            er += 1
        er += 2
    ws2.column_dimensions['A'].width = 18
    ws2.column_dimensions['B'].width = 55

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def update_timestamp(html_path):
    now = datetime.now().strftime('%d.%m.%Y %H:%M:%S')
    with open(html_path, 'r', encoding='utf-8') as f:
        html = f.read()
    html = re.sub(
        r'(id="upd-ts"[^>]*>)([^<]*)(</span>)',
        r'\g<1>Обновлено: ' + now + r'\3',
        html
    )
    with open(html_path, 'w', encoding='utf-8') as f:
        f.write(html)


# ─── HTTP Handler ────────────────────────────────────────────────────────────
class Handler(BaseHTTPRequestHandler):

    def log_message(self, format, *args):
        print(f'[{datetime.now().strftime("%H:%M:%S")}]', format % args)

    def do_GET(self):
        path = self.path.split('?')[0]

        # Генерируем xlsx на лету и отдаём браузеру
        if path == '/download':
            try:
                data = build_excel()
                from urllib.parse import quote
                fname = f'Заявочный_по_блюдам_{datetime.now().strftime("%d.%m.%Y")}.xlsx'
                fname_encoded = quote(fname, safe='')
                update_timestamp(os.path.join(BASE, 'index.html'))
                self.send_response(200)
                self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                self.send_header('Content-Disposition', f"attachment; filename*=UTF-8''{fname_encoded}")
                self.send_header('Content-Length', str(len(data)))
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(data)
            except Exception as e:
                self.send_error(500, str(e))
            return

        # Отдаём статические файлы из папки Output
        if path == '/':
            path = '/index.html'
        filepath = os.path.join(BASE, path.lstrip('/'))
        if os.path.isfile(filepath):
            ext = os.path.splitext(filepath)[1]
            ct = {'html': 'text/html; charset=utf-8', 'css': 'text/css',
                  'js': 'application/javascript'}.get(ext.lstrip('.'), 'application/octet-stream')
            with open(filepath, 'rb') as f:
                data = f.read()
            self.send_response(200)
            self.send_header('Content-Type', ct)
            self.send_header('Content-Length', str(len(data)))
            self.end_headers()
            self.wfile.write(data)
        else:
            self.send_error(404)


# ─── Запуск ──────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    PORT = 8765
    server = HTTPServer(('localhost', PORT), Handler)
    print(f'✅ Сервер запущен: http://localhost:{PORT}')
    print('   Нажми Ctrl+C чтобы остановить')
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print('\nСервер остановлен.')
